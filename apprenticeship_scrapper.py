import requests
from bs4 import BeautifulSoup
import json
from datetime import datetime
import openpyxl

PROCESSED_FILE = "processed_listings.json"

def load_processed_listings():
    """Load processed listings from a JSON file."""
    try:
        with open(PROCESSED_FILE, "r") as file:
            return set(json.load(file))
    except FileNotFoundError:
        return set()

def save_processed_listings(processed_listings):
    """Save processed listings to a JSON file."""
    with open(PROCESSED_FILE, "w") as file:
        json.dump(list(processed_listings), file)

def fetch_apprenticeships(url, processed_listings):
    """
    Fetch and parse apprenticeship listings from the given URL,
    filtering out those already processed.
    """
    try:
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8',
            'Accept-Language': 'en-US,en;q=0.5'
        }
        
        response = requests.get(url, headers=headers)
        response.raise_for_status()
        
        return parse_apprenticeships(response.text, processed_listings)
        
    except requests.RequestException as e:
        print(f"Error fetching the webpage: {e}")
        return []

def parse_apprenticeships(html_content, processed_listings):
    """
    Parse apprenticeship listings and return new ones.
    """
    soup = BeautifulSoup(html_content, 'html.parser')
    new_posts = []
    today = datetime.now()
    today_str = f"Posted {today.day} {today.strftime('%B')}"  # Creates format like "Posted 10 January"
    
    listings = soup.find_all('li', class_='das-search-results__list-item')
    
    for listing in listings:
        date_elem = listing.find('p', class_='govuk-body govuk-!-font-size-16 das-!-color-dark-grey')
        if date_elem and today_str in date_elem.text:
            title = listing.find('h2', class_='govuk-heading-m').find('a').text.strip()
            company = listing.find('p', class_='govuk-body govuk-!-margin-bottom-0').text.strip()
            location = listing.find('p', class_='govuk-body das-!-color-dark-grey').text.strip()
            
            wage_elem = listing.find('b', string='Wage')
            if wage_elem:
                wage = wage_elem.parent.get_text(strip=True).replace('Wage', '').strip()
            else:
                wage = "Not specified"
            
            closing_elem = listing.find('p', class_='govuk-body govuk-!-margin-bottom-0 govuk-!-margin-top-1')
            closing_date = closing_elem.text.strip() if closing_elem else "Not specified"
            
            job_url = listing.find('a', class_='das-search-results__link')['href']
            job_id = job_url.split("/")[-1]  # Extract unique job ID
            
            if job_id not in processed_listings:
                new_posts.append({
                    'id': job_id,
                    'title': title,
                    'location': location,
                    'company': company,
                    'wage': wage,
                    'closing_date': closing_date,
                    'job_url': f"https://www.findapprenticeship.service.gov.uk{job_url}",
                })
    
    return new_posts

def format_for_discord(apprenticeships, role_id):
    """
    Format apprenticeship listings for Discord using markdown.
    """
    if not apprenticeships:
        return f"==={role_id}===\nNo new apprenticeships found.\n"

    postings = ""
    for app in apprenticeships:
        title_with_location = f"{app['title']} in {app['location']}"
        postings += f"**{title_with_location}**\n"
        postings += f"**Wage:** {app['wage']}\n"
        postings += f"**Company:** {app['company']}\n"
        postings += f"**Closes:** {app['closing_date'].replace('Closes on ', '').replace('Closes in ', '')}\n"
        postings += f"**Link:** {app['job_url']}\n\n"
    
    return f"==={role_id}===\n{postings.strip()}\n"

def create_excel_workbook(all_listings, categories):
    """
    Create an Excel workbook with apprenticeship data.
    """
    # Create a new workbook
    wb = openpyxl.Workbook()
    
    # Remove the default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)

    # Mapping of roles to their respective category
    role_to_category = {}
    for category, data in categories.items():
        role_ids = data.get("role_ids") or [data["role_id"]]
        for role_id in role_ids:
            role_to_category[role_id] = category

    # Create sheets for each role/category
    for role_id, category in role_to_category.items():
        # Create a new sheet
        sheet = wb.create_sheet(title=role_id)
        
        # Write headers
        headers = ['Title', 'Company', 'Location', 'Wage', 'Closing Date', 'Job URL']
        for col, header in enumerate(headers, start=1):
            sheet.cell(row=1, column=col, value=header)
        
        # Filter listings for this role
        role_listings = [
            listing for listing in all_listings 
            if any(role.lower() in listing.get('title', '').lower() 
                   or role.lower() in listing.get('location', '').lower() 
                   or role.lower() in listing.get('company', '').lower() 
                   for role in [role_id])
        ]
        
        # Write listings
        for row, listing in enumerate(role_listings, start=2):
            sheet.cell(row=row, column=1, value=listing.get('title', 'N/A'))
            sheet.cell(row=row, column=2, value=listing.get('company', 'N/A'))
            sheet.cell(row=row, column=3, value=listing.get('location', 'N/A'))
            sheet.cell(row=row, column=4, value=listing.get('wage', 'N/A'))
            sheet.cell(row=row, column=5, value=listing.get('closing_date', 'N/A'))
            sheet.cell(row=row, column=6, value=listing.get('job_url', 'N/A'))
    
    # Save the workbook
    filename = f"Apprenticeships.xlsx"
    wb.save(filename)
    print(f"Excel workbook saved as {filename}")
    return filename

# Main execution
if __name__ == "__main__":
    categories = {
        "digital": {
            "url": "https://www.findapprenticeship.service.gov.uk/apprenticeships?sort=AgeAsc&searchTerm=&location=&distance=all&levelIds=6&routeIds=7",
            "role_id": "Tech",
        },
        "engineering": {
            "url": "https://www.findapprenticeship.service.gov.uk/apprenticeships?sort=AgeAsc&searchTerm=&location=&distance=all&levelIds=6&routeIds=9",
            "role_id": "Engineering",
        },
        "finance": {
            "url": "https://www.findapprenticeship.service.gov.uk/apprenticeships?sort=AgeAsc&searchTerm=&location=&distance=all&levelIds=6&routeIds=12",
            "role_ids": ["Finance", "Law"],
        },
    }
    
    processed_listings = load_processed_listings()
    all_new_listings = []
    
    for category, data in categories.items():
        url = data["url"]
        role_ids = data.get("role_ids") or [data["role_id"]]  # Handle single or multiple role IDs
        new_apprenticeships = fetch_apprenticeships(url, processed_listings)
        all_new_listings.extend(new_apprenticeships)
        
        if new_apprenticeships:
            new_ids = {app['id'] for app in new_apprenticeships}
            processed_listings.update(new_ids)
            save_processed_listings(processed_listings)
        
        for role_id in role_ids:
            discord_message = format_for_discord(new_apprenticeships, role_id)
            print(discord_message)
    
    # Create Excel workbook with all new listings
    if all_new_listings:
        create_excel_workbook(all_new_listings, categories)